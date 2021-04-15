import glob
import os
from openpyxl import load_workbook

osu_dir_base = '/mnt/md0/_datasets/OralCavity/TMA_arranged/OSU'
xls_path = '/mnt/md0/_datasets/OralCavity/TMA_arranged/OSU/OSU_Oral_Cavity_TMA_Maps.xlsx'

def lspatient(osu_dir, id2P):
    P2id = {}
    files = glob.glob(f'{osu_dir}/*.jpg')
    for f in files:
        fname = f.split('/')[-1].replace('.jpg', '')
        tmaId = fname.split('_')[1].replace('cavit', '')
        ids = fname.split('_')[2]
        id1 = ids.split('-')[0]
        id2 = ids.split('-')[1]
        id = f'{tmaId}-{id1}-{id2}'
        p = id2P[id]
        if P2id.get(id) == None:
            P2id[p] = [id]
        else:
            P2id[p].append(id)

    return P2id


def xls2dic(xls_path):
    id2P = {}
    wb = load_workbook(filename=xls_path, read_only=True)
    ws1 = wb[f'OC TMA 1']
    tmaId = '1'
    ws = ws1
    for i in range(2, ws.max_row+1):
        id1 = ws[f'D{i}'].value
        if (id1==None):
            continue
        id2 = ws[f'C{i}'].value
        patient = ws[f'E{i}'].value
        marker = ws[f'G{i}'].value
        id2P[f'{tmaId}-{id1}-{id2}'] = (patient, marker)

    for tmaId in range(2,8):
        ws = wb[f'OC TMA {tmaId}']
        for i in range(2, ws.max_row+1):
            id1 = ws[f'B{i}'].value
            if (id1==None):
                continue
            id2 = ws[f'A{i}'].value
            patient = ws[f'C{i}'].value
            marker = ws[f'E{i}'].value
            id2P[f'{tmaId}-{id1}-{id2}'] = (patient, marker)

    return id2P

if __name__ == '__main__':
    id2P = xls2dic(xls_path)
    P_masked = lspatient(f'{osu_dir_base}/masked', id2P)
#    P_nomask = lspatient(f'{osu_dir_base}/nomask', id2P)
    print('masked:', len(P_masked))
#    print('nomask:', len(P_nomask))
#    same = 0
#    for k,v in P_masked.items():
#        v2 = P_nomask.get(k)
#        if v2!=None:
#            same+=1
#    print('All:', len(P_masked)+len(P_nomask)-same)
#







